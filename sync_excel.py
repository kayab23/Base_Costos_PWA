"""
OBSOLETO: Este script ya no debe usarse en producción.

ADVERTENCIA: La lógica de la aplicación y la base de datos ya no dependen del archivo Excel.
Todos los productos, parámetros y tipos de cambio deben gestionarse directamente en la base de datos.
Este script queda solo como referencia o respaldo histórico.

Sincroniza Plantilla_Pricing_Costos_Importacion.xlsx hacia SQL Server.

Este script lee las hojas del archivo Excel de plantilla y sincroniza los datos
con las tablas correspondientes en SQL Server:

- Hoja 'Productos' → Tabla dbo.Productos (SKU, costos, origen)
- Hoja 'Parametros' → Tabla dbo.ParametrosImportacion (aranceles, fletes, etc.)
- Hoja 'TiposCambio' → Tabla dbo.TiposCambio (tasas de cambio)

El proceso:
1. Limpia completamente cada tabla antes de insertar nuevos datos
2. Lee todas las filas de cada hoja (ignorando filas vacías)
3. Inserta los datos usando fast_executemany para mejor rendimiento
4. Confirma la transacción (commit) al finalizar exitosamente

Requisitos:
- py -m pip install openpyxl pyodbc
- Driver ODBC 18 para SQL Server
- Variable de entorno SQLSERVER_CONN o ajusta DEFAULT_CONN_STR
- Archivo Excel en el mismo directorio que este script
"""
from __future__ import annotations

import os
from datetime import datetime
from decimal import Decimal
from typing import Any, Dict, Iterable, List, Sequence, Tuple

import pyodbc
from openpyxl import load_workbook

FILE_NAME = "Plantilla_Pricing_Costos_Importacion.xlsx"
DEFAULT_CONN_STR = (
    "DRIVER={ODBC Driver 18 for SQL Server};"
    "SERVER=localhost\\SQLEXPRESS;"
    "DATABASE=BD_Calculo_Costos;"
    "Trusted_Connection=yes;"
    "TrustServerCertificate=yes;"
)


def to_decimal(value: Any, default: float | None = None) -> float | None:
    if value is None:
        return default
    if isinstance(value, (int, float, Decimal)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip().replace(",", "")
        if not cleaned:
            return default
        try:
            return float(cleaned)
        except ValueError:
            return default
    return default


def bool_from_str(value: Any) -> int:
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in {"si", "sí", "yes", "y"}:
            return 1
        if normalized in {"no", "n"}:
            return 0
    return 0


# Función eliminada: register_version (tabla Versiones eliminada)


def bulk_insert(
    cursor: pyodbc.Cursor,
    table: str,
    columns: Sequence[str],
    rows: Iterable[Dict[str, Any]],
) -> None:
    rows = list(rows)
    print(f"{table}: {len(rows)} registros")
    if not rows:
        return
    placeholders = ",".join(["?"] * len(columns))
    sql = f"INSERT INTO {table} ({','.join(columns)}) VALUES ({placeholders})"
    payload: List[Tuple[Any, ...]] = [tuple(row.get(col) for col in columns) for row in rows]
    cursor.fast_executemany = True
    cursor.executemany(sql, payload)


def clear_tables(cursor: pyodbc.Cursor, tables: Sequence[str]) -> None:
    for table in tables:
        try:
            cursor.execute(f"DELETE FROM {table}")
        except pyodbc.ProgrammingError:
            # Tabla no existe, continuar
            pass


def read_catalogo(ws) -> List[Dict[str, Any]]:
    rows = []
    # Estructura CATALOGO_PRODUCTOS: SKU, Costo_Base, Descripción, Proveedor, Model, Origen, Categoría, Unidad, Moneda_Base, Fecha_Actualización, Activo (Sí/No)
    for row_data in ws.iter_rows(min_row=2, values_only=True):
        if not row_data[0]:  # sku
            continue
        sku = row_data[0]
        costo_base = row_data[1] if len(row_data) > 1 else None
        descripcion = row_data[2] if len(row_data) > 2 else ""
        proveedor = row_data[3] if len(row_data) > 3 else None
        modelo = row_data[4] if len(row_data) > 4 else None
        origen = row_data[5] if len(row_data) > 5 else None
        categoria = row_data[6] if len(row_data) > 6 else None
        unidad = row_data[7] if len(row_data) > 7 else None
        moneda = row_data[8] if len(row_data) > 8 else "USD"
        fecha_act = row_data[9] if len(row_data) > 9 else None
        activo = row_data[10] if len(row_data) > 10 else "Si"
        
        rows.append(
            {
                "sku": sku.strip() if isinstance(sku, str) else str(sku),
                "descripcion": descripcion or "",
                "proveedor": proveedor,
                "modelo": modelo,
                "origen": origen,
                "categoria": categoria,
                "unidad": unidad,
                "moneda_base": (moneda or "USD").strip() if isinstance(moneda, str) else "USD",
                "activo": bool_from_str(activo),
                "costo_base": to_decimal(costo_base),
                "fecha_actualizacion": fecha_act.date() if hasattr(fecha_act, "date") else fecha_act,
                "notas": None,
            }
        )
    return rows


# read_costos eliminada - los costos ahora están en CATALOGO_PRODUCTOS


def read_parametros(ws) -> List[Dict[str, Any]]:
    rows = []
    for row_data in ws.iter_rows(min_row=2, values_only=True):
        if not row_data or not row_data[0]:  # concepto
            continue
        concepto = row_data[0]
        tipo = row_data[1] if len(row_data) > 1 else None
        valor = row_data[2] if len(row_data) > 2 else None
        descripcion = row_data[3] if len(row_data) > 3 else None
        nota = row_data[4] if len(row_data) > 4 else None
        
        rows.append(
            {
                "concepto": concepto.strip(),
                "tipo": (tipo or "%").strip(),
                "valor": to_decimal(valor, 0.0) or 0.0,
                "descripcion": descripcion,
                "notas": nota,
                "vigente_desde": datetime.utcnow().date(),
                "vigente_hasta": None,
            }
        )
    return rows


def read_tipos_cambio(ws) -> List[Dict[str, Any]]:
    rows = []
    for moneda, tc_mxn, fuente, fecha in ws.iter_rows(min_row=2, values_only=True):
        if not moneda or not tc_mxn:
            continue
        fecha_val = fecha.date() if hasattr(fecha, "date") else datetime.utcnow().date()
        rows.append(
            {
                "moneda": moneda.strip(),
                "fecha": fecha_val,
                "tipo_cambio_mxn": to_decimal(tc_mxn, 0.0) or 0.0,
                "fuente": fuente or "Manual",
            }
        )
    return rows


# Funciones eliminadas: read_margenes y read_control_versiones (tablas eliminadas)


def main() -> None:
    conn_str = os.getenv("SQLSERVER_CONN", DEFAULT_CONN_STR)
    wb = load_workbook(FILE_NAME, data_only=True)

    sheets = {
        "catalogo": wb["CATALOGO_PRODUCTOS"],
        "parametros": wb["PARAMETROS_IMPORTACION"],
        "tipos_cambio": wb["TIPOS_CAMBIO"],
    }

    with pyodbc.connect(conn_str, autocommit=False) as conn:
        cursor = conn.cursor()

        catalog_rows = read_catalogo(sheets["catalogo"])
        param_rows = read_parametros(sheets["parametros"])
        fx_rows = read_tipos_cambio(sheets["tipos_cambio"])

        clear_tables(cursor, [
            "dbo.LandedCostCache",  # Primero eliminar tablas dependientes
            "dbo.TiposCambio",
            "dbo.ParametrosImportacion",
            "dbo.Productos",
        ])

        bulk_insert(cursor, "dbo.Productos", [
            "sku",
            "descripcion",
            "proveedor",
            "modelo",
            "origen",
            "categoria",
            "unidad",
            "moneda_base",
            "activo",
            "costo_base",
            "fecha_actualizacion",
            "notas",
        ], catalog_rows)

        bulk_insert(cursor, "dbo.ParametrosImportacion", [
            "concepto",
            "tipo",
            "valor",
            "descripcion",
            "notas",
            "vigente_desde",
            "vigente_hasta",
        ], param_rows)

        bulk_insert(cursor, "dbo.TiposCambio", [
            "moneda",
            "fecha",
            "tipo_cambio_mxn",
            "fuente",
        ], fx_rows)

        conn.commit()
        print("Sincronización completada exitosamente")


if __name__ == "__main__":
    main()
