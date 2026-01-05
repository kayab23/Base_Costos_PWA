from openpyxl import load_workbook
from pathlib import Path

FILE_NAME = "Plantilla_Pricing_Costos_Importacion.xlsx"
MAX_DATA_ROWS = 8
MAX_FORMULA_SAMPLES = 40

wb_path = Path(FILE_NAME)
if not wb_path.exists():
    raise FileNotFoundError(f"No se encontró el archivo {FILE_NAME}")

wb = load_workbook(wb_path, data_only=False)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"=== Hoja: {sheet_name} ===")
    print(f"Dimensiones: filas={ws.max_row}, columnas={ws.max_column}")

    headers = [cell.value for cell in ws[1]]
    print("Encabezados:")
    for idx, header in enumerate(headers, start=1):
        print(f"  C{idx}: {header}")

    print("Muestras de filas (hasta 8):")
    last_row = min(ws.max_row, 1 + MAX_DATA_ROWS)
    for row_idx in range(2, last_row + 1):
        row_values = []
        for col_idx in range(1, ws.max_column + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            row_values.append(value)
        print(f"  Fila {row_idx}: {row_values}")

    formula_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells.append((cell.coordinate, cell.value))

    print(f"Total de celdas con fórmulas: {len(formula_cells)}")
    for coord, formula in formula_cells[:MAX_FORMULA_SAMPLES]:
        print(f"  {coord}: {formula}")
    if len(formula_cells) > MAX_FORMULA_SAMPLES:
        print(f"  ... {len(formula_cells) - MAX_FORMULA_SAMPLES} fórmulas adicionales omitidas ...")
    print()
