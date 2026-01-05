from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
WB_NAME = "Plantilla_Pricing_Costos_Importacion.xlsx"
WB_PATH = BASE_DIR / WB_NAME
OUTPUT_PATH = BASE_DIR / "lista_output.txt"
SHEET = "LISTA_PRECIOS"

try:
    wb = load_workbook(WB_PATH, data_only=False)
    ws = wb[SHEET]

    lines: list[str] = []
    lines.append(f"Hoja {SHEET}: {ws.max_row} filas x {ws.max_column} columnas")
    headers = [cell.value for cell in ws[1]]
    lines.append("Encabezados:")
    for idx, header in enumerate(headers, start=1):
        lines.append(f"  C{idx:02d}: {header}")

    lines.append("\nFila de muestra (A2:L2):")
    row = next(ws.iter_rows(min_row=2, max_row=2, values_only=False), None)
    if row is None:
        lines.append("<sin datos>")
    else:
        values = [cell.value for cell in row]
        lines.append(str(values[:12]))
        lines.append("Formulas fila A2-L2:")
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                lines.append(f"  {cell.coordinate}: {cell.value}")

    formula_count = 0
    samples: list[tuple[str, str]] = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_count += 1
                if len(samples) < 10:
                    samples.append((cell.coordinate, cell.value))

    lines.append(f"\nTotal de celdas con fórmula: {formula_count}")
    lines.append("Formulas de ejemplo:")
    for coord, formula in samples:
        lines.append(f"  {coord}: {formula}")

    output = "\n".join(lines)
except Exception as exc:  # pragma: no cover
    import traceback

    output = f"ERROR: {exc}\n{traceback.format_exc()}"

print(output)
with open(OUTPUT_PATH, "w", encoding="utf-8") as fh:
    fh.write(output)
