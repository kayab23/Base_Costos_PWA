import sys
import json
from pathlib import Path
from openpyxl import load_workbook

DEFAULT_PATH = Path(r"C:\Users\FernandoOlveraRendon\Downloads\SKU andermed.xlsx")

def guess_column(headers, candidates):
    headers_norm = [h.strip().lower() for h in headers]
    for c in candidates:
        if c in headers_norm:
            return headers_norm.index(c)
    return None


def main(path=None):
    path = Path(path) if path else DEFAULT_PATH
    if not path.exists():
        print(json.dumps({"error": f"File not found: {path}"}))
        return 2
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active if wb is not None else None
    if ws is None or not hasattr(ws, 'iter_rows'):
        print(json.dumps({"error": "No worksheet available in workbook"}))
        return 2
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print(json.dumps({"error": "Sheet is empty"}))
        return 2
    headers = [str(c) if c is not None else "" for c in rows[0]]
    sku_candidates = ["sku", "codigo", "codigo sku", "code", "product code"]
    price_candidates = ["precio", "precio_unitario", "costo", "costo_base", "price", "price_usd", "unit price"]
    sku_idx = guess_column(headers, sku_candidates)
    price_idx = guess_column(headers, price_candidates)
    if sku_idx is None:
        # try to find any header containing 'sku' substring
        for i,h in enumerate(headers):
            if 'sku' in h.lower():
                sku_idx = i
                break
    if price_idx is None:
        for i,h in enumerate(headers):
            if any(k in h.lower() for k in ['precio','costo','price']):
                price_idx = i
                break
    results = []
    for r in rows[1:]:
        sku = None
        price = None
        if sku_idx is not None and sku_idx < len(r):
            sku = r[sku_idx]
        if price_idx is not None and price_idx < len(r):
            price = r[price_idx]
        if sku is None:
            continue
        sku_str = str(sku).strip()
        if sku_str == "":
            continue
        def safe_float(v):
            if v is None:
                return None
            try:
                # prefer numeric types directly
                if isinstance(v, (int, float)):
                    return float(v)
                s = str(v).strip()
                if s == "":
                    return None
                return float(s.replace(',', '.'))
            except Exception:
                return None

        price_val = safe_float(price)
        results.append({"sku": sku_str, "precio": price_val})
    out = {"file": str(path), "rows_total": len(rows)-1, "detected": len(results), "sample": results[:20], "items": results}
    print(json.dumps(out, ensure_ascii=False))
    return 0

if __name__ == '__main__':
    sys.exit(main(sys.argv[1] if len(sys.argv) > 1 else None))
