import sys
import json
from pathlib import Path
from openpyxl import load_workbook

DEFAULT_PATH = Path(r"C:\Users\FernandoOlveraRendon\Downloads\SKU andermed.xlsx")
OUT_SQL = Path('Scripts/updates_sku_from_excel.sql')

def guess_column(headers, candidates):
    headers_norm = [h.strip().lower() for h in headers]
    for c in candidates:
        if c in headers_norm:
            return headers_norm.index(c)
    return None


def main(path=None):
    path = Path(path) if path else DEFAULT_PATH
    if not path.exists():
        print(json.dumps({"error": f"File not found: {path}"}))
        return 2
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active if wb is not None else None
    if ws is None or not hasattr(ws, 'iter_rows'):
        print(json.dumps({"error": "No worksheet available in workbook"}))
        return 2
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print(json.dumps({"error": "Sheet is empty"}))
        return 2
    headers = [str(c) if c is not None else "" for c in rows[0]]
    sku_candidates = ["sku", "codigo", "codigo sku", "code", "product code"]
    price_candidates = ["precio", "precio_unitario", "costo", "costo_base", "price", "price_usd", "unit price"]
    sku_idx = guess_column(headers, sku_candidates)
    price_idx = guess_column(headers, price_candidates)
    if sku_idx is None:
        for i,h in enumerate(headers):
            if 'sku' in h.lower():
                sku_idx = i
                break
    if price_idx is None:
        for i,h in enumerate(headers):
            if any(k in h.lower() for k in ['precio','costo','price']):
                price_idx = i
                break
    updates = []
    for r in rows[1:]:
        if sku_idx is None:
            continue
        sku = r[sku_idx] if sku_idx < len(r) else None
        if sku is None:
            continue
        sku_str = str(sku).strip()
        if not sku_str:
            continue
        def safe_float(v):
            if v is None:
                return None
            try:
                if isinstance(v, (int, float)):
                    return float(v)
                s = str(v).strip()
                if s == "":
                    return None
                return float(s.replace(',', '.'))
            except Exception:
                return None

        price_val = None
        if price_idx is not None and price_idx < len(r):
            price = r[price_idx]
            price_val = safe_float(price)
        updates.append((sku_str, price_val))
    # Build SQL statements (update costo_base and fecha_actualizacion)
    sql_lines = [
        "-- Generated updates from SKU andermed.xlsx",
        "BEGIN TRANSACTION;",
    ]
    for sku, price in updates:
        if price is None:
            continue
        # Use parameterless literal with decimal point
        sql_lines.append(f"UPDATE dbo.Productos SET costo_base = {price}, fecha_actualizacion = GETDATE() WHERE sku = '{sku}';")
    sql_lines.append("COMMIT;")
    OUT_SQL.parent.mkdir(parents=True, exist_ok=True)
    OUT_SQL.write_text('\n'.join(sql_lines), encoding='utf-8')
    summary = {"file": str(path), "updates_count": len([u for u in updates if u[1] is not None]), "total_rows": len(rows)-1}
    print(json.dumps(summary, ensure_ascii=False))
    print(f"SQL written to {OUT_SQL}")
    return 0

if __name__ == '__main__':
    sys.exit(main(sys.argv[1] if len(sys.argv) > 1 else None))
